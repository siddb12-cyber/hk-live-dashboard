"""
Haus & Kinder — Unicommerce Data Fetcher
Runs every hour via GitHub Actions.
Fetches Sale Orders and Inventory Snapshot data.
Merges with master files and saves JSON to /data/ folder.
"""

import os
import json
import time
import requests
import csv
import io
from datetime import datetime, timedelta

# ── CONFIG ──────────────────────────────────────────────────────────────────
TENANT    = os.environ["UC_TENANT"]       # e.g. hausandkinder
USERNAME  = os.environ["UC_USERNAME"]     # Unicommerce login email
PASSWORD  = os.environ["UC_PASSWORD"]     # Unicommerce login password
CLIENT_ID = "my-trusted-client"           # Fixed for all Unicommerce accounts

BASE_URL  = f"https://{TENANT}.unicommerce.com"

# How many days of history to fetch
DAYS_BACK = 90

# ── STEP 1: GET ACCESS TOKEN ─────────────────────────────────────────────────
def get_access_token():
    print("🔐 Getting access token from Unicommerce...")
    url = (
        f"{BASE_URL}/oauth/token"
        f"?grant_type=password"
        f"&client_id={CLIENT_ID}"
        f"&username={USERNAME}"
        f"&password={PASSWORD}"
    )
    response = requests.get(url, headers={"Content-Type": "application/json"})
    if response.status_code != 200:
        raise Exception(f"Login failed: {response.status_code} — {response.text}")
    token = response.json()["access_token"]
    print("✅ Access token obtained.")
    return token


# ── STEP 2: CREATE EXPORT JOB ────────────────────────────────────────────────
def create_export_job(token, job_type, from_date, to_date):
    """
    job_type options confirmed by Unicommerce:
      - Sale Orders
      - Inventory Snapshot
    """
    print(f"📤 Creating export job: {job_type} from {from_date} to {to_date}")
    url = f"{BASE_URL}/services/rest/v1/export/createExportJob"
    headers = {
        "Authorization": f"bearer {token}",
        "Content-Type": "application/json"
    }
    payload = {
        "exportJobType": job_type,
        "searchParameters": {
            "fromDate": from_date,
            "toDate": to_date
        }
    }
    response = requests.post(url, headers=headers, json=payload)
    if response.status_code != 200:
        raise Exception(f"Export job creation failed: {response.status_code} — {response.text}")
    job_id = response.json().get("exportJobId") or response.json().get("jobId")
    print(f"✅ Export job created. ID: {job_id}")
    return job_id


# ── STEP 3: POLL UNTIL READY ─────────────────────────────────────────────────
def wait_for_export(token, job_id, max_wait_seconds=300):
    print(f"⏳ Waiting for export job {job_id} to complete...")
    url = f"{BASE_URL}/services/rest/v1/export/getExportJobStatus"
    headers = {"Authorization": f"bearer {token}"}
    for attempt in range(max_wait_seconds // 10):
        time.sleep(10)
        response = requests.get(url, headers=headers, params={"exportJobId": job_id})
        if response.status_code != 200:
            print(f"  Status check failed: {response.status_code}")
            continue
        data = response.json()
        status = data.get("status") or data.get("exportJobStatus")
        print(f"  Attempt {attempt+1}: status = {status}")
        if status in ("COMPLETED", "SUCCESS", "COMPLETE"):
            download_url = data.get("downloadUrl") or data.get("fileUrl")
            print(f"✅ Export ready. Download URL: {download_url}")
            return download_url
        elif status in ("FAILED", "ERROR"):
            raise Exception(f"Export job failed: {data}")
    raise Exception(f"Export job {job_id} timed out after {max_wait_seconds} seconds.")


# ── STEP 4: DOWNLOAD CSV ─────────────────────────────────────────────────────
def download_csv(token, download_url):
    print(f"⬇️  Downloading CSV...")
    headers = {"Authorization": f"bearer {token}"}
    response = requests.get(download_url, headers=headers)
    if response.status_code != 200:
        raise Exception(f"Download failed: {response.status_code}")
    print(f"✅ CSV downloaded. Size: {len(response.content)} bytes")
    return response.text


# ── STEP 5: PARSE CSV ────────────────────────────────────────────────────────
def parse_csv(csv_text):
    reader = csv.DictReader(io.StringIO(csv_text))
    rows = [row for row in reader]
    print(f"✅ Parsed {len(rows)} rows from CSV.")
    return rows


# ── STEP 6: LOAD MASTER FILES ────────────────────────────────────────────────
def load_masters():
    """
    Loads master Excel files from the repo root.
    Update these files in GitHub anytime — changes take effect on next run.
    """
    try:
        import openpyxl
    except ImportError:
        os.system("pip install openpyxl -q")
        import openpyxl

    masters = {}

    # --- Category Master ---
    cat_path = "Category_Wise_Inventory.xlsx"
    if os.path.exists(cat_path):
        wb = openpyxl.load_workbook(cat_path, data_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
        category_map = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            sku = str(row[0]).strip()
            record = {}
            for i, h in enumerate(headers):
                if h and i < len(row):
                    record[h] = row[i]
            category_map[sku] = {
                "category":     str(record.get("CATEGORY", "") or ""),
                "sub_category": str(record.get("SUB CATEGORY", "") or ""),
                "cogs":         float(record.get("COGS", 0) or 0),
                "weight":       float(record.get("Weight", 0) or 0),
            }
        masters["category"] = category_map
        print(f"✅ Category master loaded: {len(category_map)} SKUs")
    else:
        print("⚠️  Category master not found.")
        masters["category"] = {}

    # --- States & Cities Master ---
    state_path = "States and Cities.xlsx"
    if os.path.exists(state_path):
        wb = openpyxl.load_workbook(state_path, data_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
        pincode_map = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            try:
                pincode = str(int(row[0])).strip()
            except:
                continue
            record = {}
            for i, h in enumerate(headers):
                if h and i < len(row):
                    record[h] = row[i]
            pincode_map[pincode] = {
                "city":  str(record.get("City", "") or ""),
                "state": str(record.get("State", "") or ""),
                "zone":  str(record.get("Zone", "") or ""),
                "tier":  str(record.get("Tier Wise Classification", "") or ""),
            }
        masters["pincode"] = pincode_map
        print(f"✅ States & Cities master loaded: {len(pincode_map)} pincodes")
    else:
        print("⚠️  States master not found.")
        masters["pincode"] = {}

    # --- Channel Master ---
    ch_path = "Channel Master.xlsx"
    if os.path.exists(ch_path):
        wb = openpyxl.load_workbook(ch_path, data_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
        channel_map = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            ch_name = str(row[0]).strip()
            record = {}
            for i, h in enumerate(headers):
                if h and i < len(row):
                    record[h] = row[i]
            channel_map[ch_name] = {
                "channel_type":    str(record.get("Channel Type", "") or ""),
                "channel_abbr":    str(record.get("Channel Abbr.", "") or ""),
                "channel_display": str(record.get("Channel Abbr. + Type", "") or ""),
                "b2b_b2c":         str(record.get("Uniware Billing", "") or ""),
            }
        masters["channel"] = channel_map
        print(f"✅ Channel master loaded: {len(channel_map)} channels")
    else:
        print("⚠️  Channel master not found.")
        masters["channel"] = {}

    return masters


# ── STEP 7: ENRICH ORDERS WITH MASTER DATA ───────────────────────────────────
def enrich_orders(orders, masters):
    """Add category, channel, state info to each order row."""
    enriched = []
    for row in orders:
        # SKU info
        sku = str(row.get("SKU Code", "") or row.get("Item SKU", "") or "").strip()
        cat_info = masters["category"].get(sku, {
            "category": "Unknown", "sub_category": "Unknown", "cogs": 0, "weight": 0
        })

        # Channel info
        channel = str(row.get("Channel", "") or row.get("Facility", "") or "").strip()
        ch_info = masters["channel"].get(channel, {
            "channel_type": "Unknown", "channel_abbr": channel,
            "channel_display": channel, "b2b_b2c": "B2C"
        })

        # Pincode / state info
        pincode = str(row.get("Customer Pincode", "") or row.get("Pincode", "") or "").strip()
        try:
            pincode = str(int(float(pincode)))
        except:
            pass
        state_info = masters["pincode"].get(pincode, {
            "city":  row.get("Customer City", "Unknown"),
            "state": row.get("Customer State", "Unknown"),
            "zone":  "Unknown",
            "tier":  "Unknown"
        })

        # Order value bucket
        try:
            order_val = float(row.get("Total Amount", 0) or row.get("Sale Price", 0) or 0)
        except:
            order_val = 0

        if order_val <= 500:
            bucket = "₹0–500"
        elif order_val <= 1500:
            bucket = "₹501–1,500"
        elif order_val <= 3000:
            bucket = "₹1,501–3,000"
        elif order_val <= 6000:
            bucket = "₹3,001–6,000"
        else:
            bucket = "₹6,000+"

        enriched.append({
            **row,
            "sku":             sku,
            "category":        cat_info["category"],
            "sub_category":    cat_info["sub_category"],
            "cogs":            cat_info["cogs"],
            "weight":          cat_info["weight"],
            "channel":         channel,
            "channel_abbr":    ch_info["channel_abbr"],
            "channel_display": ch_info["channel_display"],
            "b2b_b2c":         ch_info["b2b_b2c"],
            "city":            state_info["city"],
            "state":           state_info["state"],
            "zone":            state_info["zone"],
            "tier":            state_info["tier"],
            "order_value":     order_val,
            "order_bucket":    bucket,
        })
    return enriched


# ── STEP 8: SAVE JSON ────────────────────────────────────────────────────────
def save_json(data, filename):
    os.makedirs("data", exist_ok=True)
    path = f"data/{filename}"
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, default=str)
    print(f"💾 Saved {len(data)} records to {path}")


# ── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    print("\n" + "="*60)
    print(f"🚀 HK Dashboard Fetch — {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("="*60 + "\n")

    to_date   = datetime.now().strftime("%Y-%m-%d")
    from_date = (datetime.now() - timedelta(days=DAYS_BACK)).strftime("%Y-%m-%d")
    print(f"📅 Fetching data from {from_date} to {to_date}\n")

    token = get_access_token()

    print("\n📚 Loading master files...")
    masters = load_masters()

    # Save master summaries for dashboard dropdowns
    save_json(list(masters["channel"].values()), "channels.json")
    save_json([{"sku": k, **v} for k, v in masters["category"].items()], "skus.json")

    results = {}

    # ── Sale Orders ──────────────────────────────────────────────────────────
    try:
        print("\n📦 Fetching Sale Orders...")
        job_id = create_export_job(token, "Sale Orders", from_date, to_date)
        dl_url = wait_for_export(token, job_id)
        csv_text = download_csv(token, dl_url)
        orders = parse_csv(csv_text)
        orders = enrich_orders(orders, masters)
        save_json(orders, "sale_orders.json")
        results["sale_orders"] = len(orders)
    except Exception as e:
        print(f"❌ Sale Orders failed: {e}")
        results["sale_orders"] = f"ERROR: {e}"

    # ── Inventory Snapshot ────────────────────────────────────────────────────
    try:
        print("\n🏭 Fetching Inventory Snapshot...")
        job_id = create_export_job(token, "Inventory Snapshot", from_date, to_date)
        dl_url = wait_for_export(token, job_id)
        csv_text = download_csv(token, dl_url)
        inventory = parse_csv(csv_text)
        for item in inventory:
            sku = str(item.get("SKU Code", "") or "").strip()
            cat_info = masters["category"].get(sku, {
                "category": "Unknown", "sub_category": "Unknown", "cogs": 0
            })
            item["category"]     = cat_info["category"]
            item["sub_category"] = cat_info["sub_category"]
            item["cogs"]         = cat_info["cogs"]
        save_json(inventory, "inventory.json")
        results["inventory"] = len(inventory)
    except Exception as e:
        print(f"❌ Inventory failed: {e}")
        results["inventory"] = f"ERROR: {e}"

    # ── Metadata ──────────────────────────────────────────────────────────────
    metadata = {
        "last_updated":  datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "from_date":     from_date,
        "to_date":       to_date,
        "record_counts": results
    }
    save_json(metadata, "metadata.json")

    print("\n" + "="*60)
    print("✅ FETCH COMPLETE")
    for k, v in results.items():
        print(f"   {k}: {v}")
    print("="*60 + "\n")


if __name__ == "__main__":
    main()
